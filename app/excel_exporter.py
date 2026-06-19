# -*- coding: utf-8 -*-
"""
Excel 내보내기 모듈
============================================================
수집 결과를 포맷된 Excel 파일로 출력
- 시트 구성: 전체목록 + 소스별 시트
- 헤더 고정, 필터, 조건부 서식 포함
- 클릭 가능한 URL 하이퍼링크
"""

import logging
import os
from datetime import datetime
from typing import List, Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 미설치. pip install openpyxl 실행 필요.")


# ── 스타일 상수 ──────────────────────────────────────────────────
HEADER_FILL = "1F3864"   # 네이비 (한서대 기본색)
HEADER_FONT = "FFFFFF"
ALT_ROW_FILL = "EBF1F7"  # 연한 파랑
GOLD_FILL = "C9A84C"     # 골드 (강조)
BORDER_COLOR = "BFBFBF"

# 점수별 배경색
SCORE_COLORS = {
    90: "00B050",   # 녹색: 매우 높음
    75: "92D050",   # 연녹색: 높음
    60: "FFEB9C",   # 노랑: 보통
    45: "FFC7CE",   # 연빨강: 낮음
    0:  "FFFFFF",   # 흰색: 없음
}

COLUMNS = [
    ("번호",      5),
    ("수집원",    15),
    ("기관명",    18),
    ("공고제목",  55),
    ("게시일",    12),
    ("마감일",    12),
    ("대학가능성", 10),
    ("우선순위",   10),
    ("관련부서",  20),
    ("URL",       10),
]


class ExcelExporter:
    """수집 결과 Excel 내보내기"""

    def export(
        self,
        notices: List[dict],
        output_path: str,
        title: str = "공모레이더 수집 결과",
    ) -> str:
        """
        notices: 수집된 공고 목록
        output_path: 저장 경로 (없으면 현재 디렉토리에 자동 생성)
        returns: 저장된 파일 경로
        """
        if not OPENPYXL_AVAILABLE:
            return self._export_csv_fallback(notices, output_path)

        wb = Workbook()

        # 1. 전체 목록 시트
        ws_all = wb.active
        ws_all.title = "전체목록"
        self._write_sheet(ws_all, notices, f"{title} - 전체")

        # 2. 소스별 시트
        source_groups = {}
        for n in notices:
            src = n.get("source", "기타").split("-")[0]
            source_groups.setdefault(src, []).append(n)

        for src_name, src_notices in source_groups.items():
            sheet_name = src_name[:31]  # Excel 시트명 최대 31자
            ws = wb.create_sheet(title=sheet_name)
            self._write_sheet(ws, src_notices, f"{src_name} 수집 결과")

        # 3. 요약 시트
        ws_summary = wb.create_sheet(title="요약", index=1)
        self._write_summary(ws_summary, notices, source_groups)

        # 저장
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = os.path.join(
                os.getcwd(), f"공모레이더_{timestamp}.xlsx"
            )

        os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
        wb.save(output_path)
        logger.info(f"Excel 저장 완료: {output_path}")
        return output_path

    def _write_sheet(self, ws, notices: List[dict], title: str):
        """공고 목록 시트 작성"""
        # 제목 행
        ws.merge_cells("A1:J1")
        title_cell = ws["A1"]
        title_cell.value = f"{title}  (총 {len(notices)}건, {datetime.now().strftime('%Y-%m-%d %H:%M')} 기준)"
        title_cell.font = Font(name="맑은 고딕", bold=True, size=12, color=HEADER_FONT)
        title_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # 헤더 행
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(name="맑은 고딕", bold=True, size=10, color=HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[2].height = 20

        # 데이터 행
        for row_idx, notice in enumerate(notices, start=3):
            is_alt = (row_idx % 2 == 0)
            row_fill = PatternFill("solid", fgColor=ALT_ROW_FILL) if is_alt else None

            score = notice.get("ai_score", notice.get("대학신청가능성", 0))
            score_color = self._score_color(score)

            values = [
                row_idx - 2,                                    # 번호
                notice.get("source", ""),                       # 수집원
                notice.get("agency", ""),                       # 기관명
                notice.get("title", ""),                        # 공고제목
                notice.get("post_date", ""),                    # 게시일
                notice.get("end_date", ""),                     # 마감일
                score if score else "",                         # 대학가능성
                notice.get("priority", notice.get("우선순위", "")),  # 우선순위
                notice.get("departments", notice.get("관련부서", "")),  # 관련부서
                "링크",                                          # URL
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name="맑은 고딕", size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))
                cell.border = self._border()

                if row_fill and col_idx != 7:
                    cell.fill = row_fill

                # 점수 열 색상
                if col_idx == 7 and score:
                    cell.fill = PatternFill("solid", fgColor=score_color)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # URL 하이퍼링크
                if col_idx == 10:
                    url = notice.get("url", "")
                    if url:
                        cell.hyperlink = url
                        cell.font = Font(
                            name="맑은 고딕", size=9,
                            color="0563C1", underline="single"
                        )

            ws.row_dimensions[row_idx].height = 30

        # 헤더 고정 및 필터
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:J{len(notices) + 2}"

    def _write_summary(self, ws, all_notices: List[dict], groups: dict):
        """요약 시트 작성"""
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20

        # 제목
        ws.merge_cells("A1:C1")
        ws["A1"].value = f"공모레이더 수집 요약  ({datetime.now().strftime('%Y-%m-%d %H:%M')})"
        ws["A1"].font = Font(name="맑은 고딕", bold=True, size=13, color=HEADER_FONT)
        ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # 전체 통계
        row = 3
        stats = [
            ("전체 수집 건수", len(all_notices), "건"),
            ("대학 신청 가능(70점+)", sum(1 for n in all_notices if n.get("ai_score", 0) >= 70), "건"),
            ("AI 분석 완료", sum(1 for n in all_notices if n.get("ai_score")), "건"),
            ("수집 기관 수", len(groups), "개"),
        ]
        for label, value, unit in stats:
            ws.cell(row=row, column=1, value=label).font = Font(name="맑은 고딕", bold=True, size=10)
            ws.cell(row=row, column=2, value=f"{value:,} {unit}").font = Font(name="맑은 고딕", size=10)
            ws.row_dimensions[row].height = 20
            row += 1

        # 소스별 통계
        row += 1
        headers = ["수집원", "건수", "비고"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = Font(name="맑은 고딕", bold=True, color=HEADER_FONT)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for src, src_notices in sorted(groups.items(), key=lambda x: -len(x[1])):
            ws.cell(row=row, column=1, value=src).font = Font(name="맑은 고딕", size=9)
            ws.cell(row=row, column=2, value=len(src_notices)).font = Font(name="맑은 고딕", size=9)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            if row % 2 == 0:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=ALT_ROW_FILL)
            row += 1

    def _border(self):
        side = Side(style="thin", color=BORDER_COLOR)
        return Border(left=side, right=side, top=side, bottom=side)

    def _score_color(self, score) -> str:
        try:
            s = int(score)
        except (TypeError, ValueError):
            return "FFFFFF"
        for threshold in sorted(SCORE_COLORS.keys(), reverse=True):
            if s >= threshold:
                return SCORE_COLORS[threshold]
        return "FFFFFF"

    def _export_csv_fallback(self, notices: List[dict], output_path: str) -> str:
        """openpyxl 없을 때 CSV로 대체 저장"""
        import csv
        if not output_path:
            output_path = f"공모레이더_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        csv_path = output_path.replace(".xlsx", ".csv")

        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["번호", "수집원", "기관명", "공고제목", "게시일", "마감일",
                              "대학가능성", "관련부서", "URL"])
            for i, n in enumerate(notices, 1):
                writer.writerow([
                    i,
                    n.get("source", ""),
                    n.get("agency", ""),
                    n.get("title", ""),
                    n.get("post_date", ""),
                    n.get("end_date", ""),
                    n.get("ai_score", ""),
                    n.get("departments", ""),
                    n.get("url", ""),
                ])

        logger.info(f"CSV 저장 완료 (openpyxl 없음): {csv_path}")
        return csv_path
